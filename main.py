import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
import os
from openpyxl import load_workbook
from datetime import datetime
from icalendar import Calendar, Event
from ttkthemes import ThemedTk

class EventExtractorApp(ThemedTk):
    def __init__(self):
        super().__init__(theme="adapta")
        self.title("iCal Event Extractor")
        self.geometry("700x500")
        self.minsize(500, 300)
        self.events = []
        self.create_menu()
        self.create_widgets()
        self.create_status_bar()

    def create_menu(self):
        menubar = tk.Menu(self)
        theme_menu = tk.Menu(menubar, tearoff=0)
        self.available_themes = [
            "arc", "plastik", "clearlooks", "radiance", "scidgrey", "scidgreen", "scidmint", "scidblue", "scidpurple", "scidpink", "scidgrey", "black", "breeze", "equilux", "keramik", "winxpblue", "yaru", "adapta", "alt", "clam", "default", "classic"
        ]
        for theme in self.available_themes:
            theme_menu.add_command(label=theme, command=lambda t=theme: self.apply_theme(t))
        menubar.add_cascade(label="Theme", menu=theme_menu)
        self.config(menu=menubar)

    def apply_theme(self, theme_name):
        try:
            super().set_theme(theme_name)
        except Exception as e:
            messagebox.showerror("Theme Error", f"Could not set theme '{theme_name}': {e}")

    def create_widgets(self):
        # Main frame for layout
        main_frame = tk.Frame(self)
        main_frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        # Set ttk theme to 'clam' for better cross-platform Treeview visibility
        style = ttk.Style()
        try:
            style.theme_use('clam')
        except Exception:
            pass  # fallback to default if 'clam' is not available
        style.configure("Treeview", rowheight=24, borderwidth=2, relief="groove")
        style.configure("Treeview.Heading", font=(None, 10, "bold"))
        style.map("Treeview", background=[('selected', '#ececec')])

        # Import button
        self.import_btn = tk.Button(main_frame, text="Import XLSM or PDF", command=self.import_file)
        self.import_btn.pack(pady=(0, 10), anchor="w")

        # Search/filter bar
        search_frame = tk.Frame(main_frame)
        search_frame.pack(fill=tk.X, pady=(0, 5))
        tk.Label(search_frame, text="Search:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_var.trace_add('write', lambda *args: self.refresh_tree())
        self.search_entry = tk.Entry(search_frame, textvariable=self.search_var)
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))

        # Label above the table
        self.table_label = tk.Label(main_frame, text="Extracted Events:")
        self.table_label.pack(anchor="w")

        # Treeview with vertical scrollbar
        columns = ("date", "event_type", "project_code", "notes")
        tree_frame = tk.Frame(main_frame)
        tree_frame.pack(expand=True, fill=tk.BOTH)
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="extended")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)
        # Modern font and row height
        style = ttk.Style()
        style.configure("Treeview", font=("Segoe UI", 11), rowheight=28)
        style.configure("Treeview.Heading", font=("Segoe UI", 11, "bold"))
        # Alternating row colors
        style.map("Treeview", background=[('selected', '#3399FF')])
        self.tree.tag_configure('oddrow', background='#f5f5f5')
        self.tree.tag_configure('evenrow', background='#e0e0e0')
        # Resizable columns (default in ttk, but set minwidth for usability)
        for i, col in enumerate(columns):
            self.tree.heading(col, text=col.capitalize(), command=lambda c=col: self.sort_by_column(c, False))
            self.tree.column(col, width=150, minwidth=80, stretch=True)
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        self.tree.bind("<Button-3>", self.on_tree_right_click)
        self.tree.bind("<Button-2>", self.on_tree_right_click)

        # Export button
        self.export_btn = tk.Button(main_frame, text="Export to .ics", command=self.export_ics, state=tk.DISABLED)
        self.export_btn.pack(pady=10, anchor="e")
        self.refresh_tree()

    def create_status_bar(self):
        self.status_var = tk.StringVar()
        self.status_var.set("Ready.")
        self.status_bar = tk.Label(self, textvariable=self.status_var, bd=1, relief=tk.SUNKEN, anchor="w")
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def set_status(self, message, error=False):
        self.status_var.set(message)
        if error:
            self.status_bar.config(fg="red")
        else:
            self.status_bar.config(fg="black")

    def import_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Macro-Enabled Workbook", "*.xlsm"), ("PDF Files", "*.pdf")]
        )
        if not file_path:
            return
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".xlsm":
            self.parse_xlsm(file_path)
        elif ext == ".pdf":
            self.parse_pdf(file_path)
        else:
            messagebox.showerror("Invalid file", "Please select an XLSM or PDF file.")

    def parse_xlsm(self, file_path):
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            # Only treat these as event columns
            allowed_event_columns = {"Zustellung zu CCR", "CCR", "Zustellung zu ITV"}
            event_columns = []  # List of (index, column_name)
            for i, h in enumerate(headers):
                if h and str(h).strip() in allowed_event_columns:
                    event_columns.append((i, str(h).strip()))
            events = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Find the first datetime value in the row as the date
                date_val = None
                for cell in row:
                    if isinstance(cell, datetime):
                        date_val = cell
                        break
                if not date_val:
                    continue  # Skip if no valid date found
                date_str = date_val.strftime("%Y-%m-%d")
                # For each allowed event column that is non-empty, create an event
                for idx, col_name in event_columns:
                    project_code = row[idx]
                    if project_code and str(project_code).strip():
                        events.append({
                            "date": date_str,
                            "event_type": col_name,
                            "project_code": str(project_code).strip(),
                            "notes": ""
                        })
            self.events = events
            print(f"[DEBUG] Parsed {len(self.events)} events.")
            self.refresh_tree()
            if self.events:
                self.export_btn.config(state=tk.NORMAL)
            else:
                self.export_btn.config(state=tk.DISABLED)
            self.set_status(f"Imported {len(self.events)} events from XLSM.")
        except Exception as e:
            self.set_status(f"Failed to parse XLSM: {e}", error=True)
            messagebox.showerror("Error", f"Failed to parse XLSM: {e}")

    def parse_pdf(self, file_path):
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                # Try to find the first page with a table
                table = None
                for page in pdf.pages:
                    tables = page.extract_tables()
                    if tables:
                        table = tables[0]
                        break
                if not table:
                    self.set_status("No table found in PDF.", error=True)
                    messagebox.showerror("Error", "No table found in PDF.")
                    return
                headers = table[0]
                # Normalize header names (remove whitespace and line breaks)
                def norm_header(h):
                    if not h:
                        return None
                    return ' '.join(str(h).replace('\n', ' ').split())
                norm_headers = [norm_header(h) for h in headers]
                allowed_event_columns = {"Zustellung zu CCR", "CCR", "Zustellung zu ITV"}
                event_columns = []
                for i, h in enumerate(norm_headers):
                    if h and h in allowed_event_columns:
                        event_columns.append((i, h))
                events = []
                for row in table[1:]:
                    # Find the first cell that looks like a date (YYYY-MM-DD or DD.MM.YYYY)
                    date_val = None
                    for cell in row:
                        if cell:
                            cell_str = cell.strip()
                            for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
                                try:
                                    date_val = datetime.strptime(cell_str, fmt)
                                    break
                                except Exception:
                                    continue
                            if date_val:
                                break
                    if not date_val:
                        continue
                    date_str = date_val.strftime("%Y-%m-%d")
                    for idx, col_name in event_columns:
                        if idx < len(row):
                            project_code = row[idx]
                            if project_code and str(project_code).strip():
                                events.append({
                                    "date": date_str,
                                    "event_type": col_name,
                                    "project_code": str(project_code).strip(),
                                    "notes": ""
                                })
                self.events = events
                self.refresh_tree()
                if self.events:
                    self.export_btn.config(state=tk.NORMAL)
                else:
                    self.export_btn.config(state=tk.DISABLED)
                self.set_status(f"Imported {len(self.events)} events from PDF.")
        except Exception as e:
            self.set_status(f"Failed to parse PDF: {e}", error=True)
            messagebox.showerror("Error", f"Failed to parse PDF: {e}")

    def refresh_tree(self):
        print(f"[DEBUG] refresh_tree called. Number of events: {len(self.events)}")
        for i in self.tree.get_children():
            self.tree.delete(i)
        # Filter events based on search
        search = getattr(self, 'search_var', None)
        if search:
            search_text = self.search_var.get().strip().lower()
        else:
            search_text = ''
        filtered_events = []
        if search_text:
            for event in self.events:
                if (search_text in event["date"].lower() or
                    search_text in event["event_type"].lower() or
                    search_text in event["project_code"].lower() or
                    search_text in event["notes"].lower()):
                    filtered_events.append(event)
        else:
            filtered_events = self.events
        if not filtered_events:
            self.tree.insert("", tk.END, iid=0, values=("No events loaded", "", "", ""))
            self.tree.item(0, tags=("placeholder",))
            self.tree.tag_configure("placeholder", foreground="#888888")
        else:
            for idx, event in enumerate(filtered_events):
                tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
                self.tree.insert("", tk.END, iid=idx, values=(event["date"], event["event_type"], event["project_code"], event["notes"]), tags=(tag,))

    def on_tree_double_click(self, event):
        item_id = self.tree.focus()
        if not item_id:
            return
        idx = int(item_id)
        event_data = self.events[idx]
        # Identify which column was clicked
        region = self.tree.identify("region", event.x, event.y)
        col = self.tree.identify_column(event.x)
        col_num = int(col.replace('#', '')) - 1  # columns are 1-indexed
        columns = ("date", "event_type", "project_code", "notes")
        if region == "cell" and columns[col_num] == "notes":
            # Inline edit for notes
            self.inline_edit_notes(idx)
        else:
            # Popup edit for all fields
            EditEventDialog(self, event_data, lambda updated_event: self.update_event(idx, updated_event))

    def on_tree_right_click(self, event):
        # Select the row under the cursor if not already selected
        row_id = self.tree.identify_row(event.y)
        if row_id:
            if row_id not in self.tree.selection():
                self.tree.selection_set(row_id)
        selected = self.tree.selection()
        if not selected:
            return
        menu = tk.Menu(self, tearoff=0)
        if len(selected) == 1:
            menu.add_command(label="Edit", command=lambda: self.edit_selected_event(selected[0]))
            menu.add_command(label="Delete", command=lambda: self.delete_selected_event(selected[0]))
        menu.add_command(label="Bulk Edit Notes", command=self.bulk_edit_notes)
        menu.add_command(label="Bulk Delete", command=self.bulk_delete_events)
        menu.tk_popup(event.x_root, event.y_root)

    def edit_selected_event(self, iid):
        idx = int(iid)
        event_data = self.events[idx]
        EditEventDialog(self, event_data, lambda updated_event: self.update_event(idx, updated_event))

    def delete_selected_event(self, iid):
        idx = int(iid)
        del self.events[idx]
        self.refresh_tree()
        self.set_status("Event deleted.")

    def bulk_delete_events(self):
        selected = self.tree.selection()
        if not selected:
            self.set_status("No events selected for bulk delete.", error=True)
            return
        idxs = sorted([int(iid) for iid in selected], reverse=True)
        for idx in idxs:
            del self.events[idx]
        self.refresh_tree()
        self.set_status(f"Deleted {len(idxs)} events.")

    def bulk_edit_notes(self):
        selected = self.tree.selection()
        if not selected:
            self.set_status("No events selected for bulk edit.", error=True)
            return
        idxs = [int(iid) for iid in selected]
        # Prompt for new note
        new_note = simpledialog.askstring("Bulk Edit Notes", f"Enter new note for {len(idxs)} selected events:")
        if new_note is not None:
            for idx in idxs:
                self.events[idx]["notes"] = new_note
            self.refresh_tree()
            self.set_status(f"Updated notes for {len(idxs)} events.")

    def inline_edit_notes(self, idx):
        item_id = str(idx)
        x, y, width, height = self.tree.bbox(item_id, 3)  # 3 = notes column (0-based)
        value = self.events[idx]["notes"]
        entry = tk.Entry(self.tree)
        entry.insert(0, value)
        entry.select_range(0, tk.END)
        entry.focus()
        entry.place(x=x, y=y, width=width, height=height)

        def save_edit(event=None):
            new_val = entry.get()
            self.events[idx]["notes"] = new_val
            entry.destroy()
            self.refresh_tree()
            self.set_status(f"Note updated for event on {self.events[idx]['date']}.")

        entry.bind("<Return>", save_edit)
        entry.bind("<FocusOut>", lambda e: save_edit())

    def update_event(self, idx, updated_event):
        self.events[idx] = updated_event
        self.refresh_tree()

    def export_ics(self):
        if not self.events:
            self.set_status("There are no events to export.", error=True)
            messagebox.showwarning("No Events", "There are no events to export.")
            return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".ics",
            filetypes=[("iCal files", "*.ics")],
            title="Save iCal File"
        )
        if not file_path:
            self.set_status("Export cancelled.")
            return
        try:
            cal = Calendar()
            cal.add('prodid', '-//iCal Event Extractor//mxm.dk//')
            cal.add('version', '2.0')
            for event in self.events:
                try:
                    event_date = datetime.strptime(event['date'], "%Y-%m-%d").date()
                except Exception:
                    continue  # Skip events with invalid dates
                ical_event = Event()
                ical_event.add('summary', f"{event['event_type']}: {event['project_code']}")
                ical_event.add('dtstart', event_date)
                ical_event.add('dtend', event_date)
                ical_event.add('description', event['notes'])
                ical_event.add('transp', 'TRANSPARENT')
                ical_event.add('X-MICROSOFT-CDO-ALLDAYEVENT', 'TRUE')
                cal.add_component(ical_event)
            with open(file_path, 'wb') as f:
                f.write(cal.to_ical())
            self.set_status(f"ICS file saved to: {file_path}")
            messagebox.showinfo("Export Successful", f"ICS file saved to: {file_path}")
        except Exception as e:
            self.set_status(f"Failed to export ICS: {e}", error=True)
            messagebox.showerror("Export Failed", f"Failed to export ICS: {e}")

    def sort_by_column(self, col, reverse):
        # Get column index
        columns = ("date", "event_type", "project_code", "notes")
        col_idx = columns.index(col)
        # Get all items and sort
        data = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        try:
            data.sort(key=lambda t: t[0], reverse=reverse)
        except Exception:
            data.sort(key=lambda t: str(t[0]), reverse=reverse)
        # Rearrange items in sorted positions
        for index, (val, k) in enumerate(data):
            self.tree.move(k, '', index)
        # Reverse sort next time
        self.tree.heading(col, command=lambda: self.sort_by_column(col, not reverse))

class EditEventDialog(tk.Toplevel):
    def __init__(self, parent, event_data, on_save):
        super().__init__(parent)
        self.title("Edit Event")
        self.resizable(False, False)
        self.event_data = event_data.copy()
        self.on_save = on_save
        self.result = None
        self.grab_set()

        # Date
        tk.Label(self, text="Date (YYYY-MM-DD):").grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.date_var = tk.StringVar(value=event_data["date"])
        self.date_entry = tk.Entry(self, textvariable=self.date_var)
        self.date_entry.grid(row=0, column=1, padx=5, pady=5)

        # Event Type
        tk.Label(self, text="Event Type:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.event_type_var = tk.StringVar(value=event_data["event_type"])
        self.event_type_entry = tk.Entry(self, textvariable=self.event_type_var)
        self.event_type_entry.grid(row=1, column=1, padx=5, pady=5)

        # Project Code
        tk.Label(self, text="Project Code:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
        self.project_code_var = tk.StringVar(value=event_data["project_code"])
        self.project_code_entry = tk.Entry(self, textvariable=self.project_code_var)
        self.project_code_entry.grid(row=2, column=1, padx=5, pady=5)

        # Notes
        tk.Label(self, text="Notes:").grid(row=3, column=0, sticky="e", padx=5, pady=5)
        self.notes_var = tk.StringVar(value=event_data["notes"])
        self.notes_entry = tk.Entry(self, textvariable=self.notes_var)
        self.notes_entry.grid(row=3, column=1, padx=5, pady=5)

        # Buttons
        btn_frame = tk.Frame(self)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=10)
        save_btn = tk.Button(btn_frame, text="Save", command=self.save)
        save_btn.pack(side=tk.LEFT, padx=5)
        cancel_btn = tk.Button(btn_frame, text="Cancel", command=self.destroy)
        cancel_btn.pack(side=tk.LEFT, padx=5)

    def save(self):
        # Validate date
        date_str = self.date_var.get().strip()
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
        except Exception:
            messagebox.showerror("Invalid Date", "Date must be in YYYY-MM-DD format.")
            return
        updated_event = {
            "date": date_str,
            "event_type": self.event_type_var.get().strip(),
            "project_code": self.project_code_var.get().strip(),
            "notes": self.notes_var.get().strip(),
        }
        self.on_save(updated_event)
        self.destroy()

if __name__ == "__main__":
    app = EventExtractorApp()
    app.mainloop() 